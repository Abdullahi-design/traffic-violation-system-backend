import io
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook


def generate_pdf_report(title: str, headers: list[str], rows: list[list], summary: dict | None = None) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 0.3 * inch))

    if summary:
        for key, value in summary.items():
            elements.append(Paragraph(f"<b>{key}:</b> {value}", styles["Normal"]))
        elements.append(Spacer(1, 0.3 * inch))

    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
    ]))
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(title: str, headers: list[str], rows: list[list], summary: dict | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    row_num = 1
    if summary:
        for key, value in summary.items():
            ws.cell(row=row_num, column=1, value=key)
            ws.cell(row=row_num, column=2, value=str(value))
            row_num += 1
        row_num += 1

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    for row_data in rows:
        row_num += 1
        for col_idx, value in enumerate(row_data, 1):
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_num, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
